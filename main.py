from fastapi import FastAPI, Request, HTTPException
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel
from typing import Dict, Any, Optional
import pandas as pd
import asyncio
import os
import re
import shutil
import tempfile
import io
from datetime import datetime, date
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import threading
import time
import subprocess
import sys
from openpyxl import load_workbook

app = FastAPI()
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

# ===== CONFIGURATION =====
FILE_PATHS = [
    "mockData.xlsx",
    # Add more file paths here:
    # "//server/shared/projects.xlsx",
    # "another.xlsx",
]

# ===== GLOBAL CACHE =====
cached_data = {
    "all_sheets_data": {},
    "sheet_names": [],
    "last_updated": None
}
data_version = 0  # Increments when data changes
connected_clients = []  # SSE clients waiting for updates

# ===== WRITE LOCK (prevents Watchdog reload during saves) =====
write_in_progress = False
write_lock = threading.Lock()

# ===== EXCEL PROCESS TRACKING =====
excel_processes = {}  # abs_file_path -> subprocess.Popen

# ===== FILE WATCHER =====
class ExcelFileHandler(FileSystemEventHandler):
    """Watches for changes to Excel files and triggers data reload."""

    def __init__(self, file_paths):
        self.file_paths = [os.path.abspath(fp) for fp in file_paths]
        self.last_reload = 0
        self.debounce_seconds = 3  # Wait 3 seconds before reloading to let Excel finish saving

    def on_modified(self, event):
        global write_in_progress

        if event.is_directory:
            return

        # Check if the modified file is one we're watching
        modified_path = os.path.abspath(event.src_path)

        # Only watch .xlsx files
        if not modified_path.lower().endswith('.xlsx'):
            return

        # Excel creates temp files like ~$filename.xlsx, ignore those
        if os.path.basename(modified_path).startswith('~$'):
            return

        # Only react to files explicitly listed in FILE_PATHS
        if modified_path not in self.file_paths:
            return

        # Skip if we're currently writing to prevent reload loops
        with write_lock:
            if write_in_progress:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Ignoring change during write operation")
                return

        current_time = time.time()
        # Debounce: only reload if enough time has passed
        if current_time - self.last_reload > self.debounce_seconds:
            self.last_reload = current_time
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Detected change in: {os.path.basename(modified_path)}")
            # Wait a moment for Excel to finish writing before reading
            time.sleep(0.5)
            reload_data()

def start_file_watcher():
    """Start watching all Excel file directories for changes."""
    global observer

    event_handler = ExcelFileHandler(FILE_PATHS)
    observer = Observer()

    # Watch the directories containing the Excel files
    watched_dirs = set()
    for file_path in FILE_PATHS:
        abs_path = os.path.abspath(file_path)
        dir_path = os.path.dirname(abs_path) or '.'
        if dir_path not in watched_dirs:
            watched_dirs.add(dir_path)
            observer.schedule(event_handler, dir_path, recursive=False)
            print(f"Watching directory: {dir_path}")

    observer.start()
    print("File watcher started - will auto-refresh when Excel files change")

def reload_data(max_retries=3, retry_delay=1.0):
    """Reload data from Excel files and notify connected clients.

    Includes retry logic to handle cases where Excel is still saving
    and the file might be temporarily incomplete or unavailable.
    """
    global data_version

    for attempt in range(max_retries):
        try:
            all_sheets_data, sheet_names = load_all_sheets_data()

            # Validate we got real data (not just defaults due to read failure)
            has_real_data = False
            for sheet_name, tasks in all_sheets_data.items():
                if sheet_name != 'Default':
                    for task_name, instances in tasks.items():
                        if task_name != 'Sample Task':
                            has_real_data = True
                            break
                if has_real_data:
                    break

            # If we only have default data but we had real data before, retry
            if not has_real_data and cached_data["sheet_names"] and 'Default' not in cached_data["sheet_names"]:
                if attempt < max_retries - 1:
                    print(f"[{datetime.now().strftime('%H:%M:%S')}] Got empty data, retrying in {retry_delay}s... (attempt {attempt + 1}/{max_retries})")
                    time.sleep(retry_delay)
                    continue
                else:
                    print(f"[{datetime.now().strftime('%H:%M:%S')}] Warning: Could not read valid data after {max_retries} attempts, keeping previous data")
                    return  # Keep existing cached data

            # Data looks valid, update cache
            cached_data["all_sheets_data"] = all_sheets_data
            cached_data["sheet_names"] = sheet_names
            cached_data["last_updated"] = datetime.now().isoformat()
            data_version += 1

            print(f"[{datetime.now().strftime('%H:%M:%S')}] Data reloaded (version {data_version})")

            # Notify all connected SSE clients
            notify_clients()
            return

        except Exception as e:
            if attempt < max_retries - 1:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Error reloading data: {e}, retrying in {retry_delay}s... (attempt {attempt + 1}/{max_retries})")
                time.sleep(retry_delay)
            else:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Error reloading data after {max_retries} attempts: {e}")

def read_file_with_shared_access(file_path: str) -> bytes:
    """
    Read a file even if it's open in another program (like Excel).
    Uses shared read access on Windows.
    """
    import sys
    if sys.platform == 'win32':
        # Windows: use win32 API for shared read access
        import msvcrt
        import ctypes
        from ctypes import wintypes

        GENERIC_READ = 0x80000000
        FILE_SHARE_READ = 0x1
        FILE_SHARE_WRITE = 0x2
        FILE_SHARE_DELETE = 0x4
        OPEN_EXISTING = 3
        FILE_ATTRIBUTE_NORMAL = 0x80

        CreateFileW = ctypes.windll.kernel32.CreateFileW
        CreateFileW.argtypes = [wintypes.LPCWSTR, wintypes.DWORD, wintypes.DWORD,
                                wintypes.LPVOID, wintypes.DWORD, wintypes.DWORD, wintypes.HANDLE]
        CreateFileW.restype = wintypes.HANDLE

        ReadFile = ctypes.windll.kernel32.ReadFile
        GetFileSize = ctypes.windll.kernel32.GetFileSize
        CloseHandle = ctypes.windll.kernel32.CloseHandle

        handle = CreateFileW(
            file_path,
            GENERIC_READ,
            FILE_SHARE_READ | FILE_SHARE_WRITE | FILE_SHARE_DELETE,
            None,
            OPEN_EXISTING,
            FILE_ATTRIBUTE_NORMAL,
            None
        )

        if handle == -1:
            raise IOError(f"Cannot open file: {file_path}")

        try:
            file_size = GetFileSize(handle, None)
            buffer = ctypes.create_string_buffer(file_size)
            bytes_read = wintypes.DWORD()
            ReadFile(handle, buffer, file_size, ctypes.byref(bytes_read), None)
            return buffer.raw[:bytes_read.value]
        finally:
            CloseHandle(handle)
    else:
        # Unix/Mac: regular read usually works
        with open(file_path, 'rb') as f:
            return f.read()


def safe_read_excel(file_path: str, **kwargs):
    """
    Safely read an Excel file that might be open in Excel.
    Reads file content with shared access, then loads from memory.
    """
    file_bytes = read_file_with_shared_access(file_path)
    return pd.read_excel(io.BytesIO(file_bytes), **kwargs)


def safe_get_sheet_names(file_path: str) -> list:
    """
    Safely get sheet names from an Excel file that might be open in Excel.
    """
    file_bytes = read_file_with_shared_access(file_path)
    with pd.ExcelFile(io.BytesIO(file_bytes)) as excel_file:
        return excel_file.sheet_names


def load_all_sheets_data():
    all_data = {}
    valid_sheet_names = []

    for file_path in FILE_PATHS:
        abs_file_path = os.path.abspath(file_path)  # Store absolute path for metadata

        # Retry logic for getting sheet names (file might be mid-save)
        sheet_names = None
        for attempt in range(3):
            try:
                sheet_names = safe_get_sheet_names(file_path)
                break
            except Exception as e:
                if attempt < 2:
                    time.sleep(0.3)  # Brief wait before retry
                else:
                    print(f"Error opening file '{file_path}' after 3 attempts: {e}")

        if sheet_names is None:
            continue

        for sheet_name in sheet_names:
            # Skip default Excel sheet names like "Sheet1", "Sheet2", etc.
            if re.match(r'^sheet\d+$', sheet_name.lower().strip()):
                print(f"Skipping default sheet name: '{sheet_name}'")
                continue

            try:
                # Retry logic for reading sheet data
                df = None
                for read_attempt in range(3):
                    try:
                        df = safe_read_excel(file_path, sheet_name=sheet_name)
                        break
                    except Exception as read_err:
                        if read_attempt < 2:
                            time.sleep(0.3)
                        else:
                            raise read_err

                if df is None or df.empty or len(df.columns) < 2:
                    print(f"Warning: Skipping empty sheet '{sheet_name}' in '{file_path}'")
                    continue

                # Rule 1: Stop at the first unnamed/undefined column
                valid_cols = []
                for col in df.columns:
                    col_str = str(col).strip()
                    if col_str.startswith('Unnamed') or col_str == '' or col_str == 'nan':
                        break
                    valid_cols.append(col)

                if len(valid_cols) < 2:
                    print(f"Warning: Skipping sheet '{sheet_name}' in '{file_path}' - not enough named columns")
                    continue

                # Store all column names for metadata
                all_columns = [str(col) for col in valid_cols]

                df = df[valid_cols]

                # Rule 2: Stop at the first empty Task Name (first column) row
                task_name_col = df.columns[0]
                cut_index = None
                for i, value in enumerate(df[task_name_col]):
                    if pd.isna(value) or str(value).strip() == '':
                        cut_index = i
                        break

                if cut_index is not None:
                    df = df.iloc[:cut_index]

                if df.empty:
                    print(f"Warning: Skipping sheet '{sheet_name}' in '{file_path}' - no valid data")
                    continue

                # Initialize sheet if it doesn't exist
                if sheet_name not in all_data:
                    all_data[sheet_name] = {}
                    valid_sheet_names.append(sheet_name)

                # Use enumerate to track row index for metadata
                for row_idx, (_, row) in enumerate(df.iterrows()):
                    task_name = str(row[task_name_col])

                    if task_name == 'nan' or not task_name.strip():
                        continue

                    # Build details string and raw values dict
                    details = []
                    raw_values = {}
                    for col in df.columns:
                        value = row[col]
                        # Store raw value (None if NaN), convert Timestamps to ISO strings
                        if pd.isna(value):
                            raw_values[str(col)] = None
                        elif isinstance(value, (pd.Timestamp, datetime, date)):
                            raw_values[str(col)] = value.isoformat()
                        else:
                            raw_values[str(col)] = value
                        # Add to details string (skip task name column and empty values)
                        if col != task_name_col and pd.notna(value) and str(value).strip():
                            details.append(f"{col}: {value}")

                    formatted_details = '\n'.join(details)

                    # Create enriched entry with metadata
                    entry = {
                        "details": formatted_details,
                        "metadata": {
                            "file_path": abs_file_path,
                            "sheet_name": sheet_name,
                            "row_index": row_idx,
                            "columns": all_columns,
                            "raw_values": raw_values,
                            "task_name": task_name  # For row validation during save
                        }
                    }

                    # Merge: append to existing task or create new
                    if task_name not in all_data[sheet_name]:
                        all_data[sheet_name][task_name] = []
                    all_data[sheet_name][task_name].append(entry)

                print(f"Loaded sheet '{sheet_name}' from '{file_path}'")

            except Exception as e:
                print(f"Error loading sheet '{sheet_name}' from '{file_path}': {e}")
                continue

    if not valid_sheet_names:
        print("Warning: No valid sheets found, creating default")
        all_data['Default'] = {
            'Sample Task': [{
                "details": "No data available\nPlease check your Excel file",
                "metadata": None
            }]
        }
        valid_sheet_names = ['Default']

    print(f"Total sheets loaded: {len(valid_sheet_names)}")
    return all_data, valid_sheet_names


# ===== SSE (Server-Sent Events) for real-time updates =====
def notify_clients():
    """Send update notification to all connected SSE clients."""
    global connected_clients
    # Mark all clients as needing an update
    for client in connected_clients:
        client["needs_update"] = True


async def event_generator():
    """Generate SSE events for a connected client."""
    global data_version

    client = {"needs_update": False, "last_version": data_version}
    connected_clients.append(client)

    try:
        while True:
            # Check if there's a new version
            if client["needs_update"] or client["last_version"] != data_version:
                client["needs_update"] = False
                client["last_version"] = data_version
                yield f"data: {data_version}\n\n"

            await asyncio.sleep(0.5)  # Check every 500ms
    finally:
        connected_clients.remove(client)


@app.get("/events")
async def sse_events():
    """SSE endpoint for real-time data update notifications."""
    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )


@app.get("/api/data")
async def get_data():
    """API endpoint to fetch the latest cached data."""
    return {
        "all_sheets_data": cached_data["all_sheets_data"],
        "sheet_names": cached_data["sheet_names"],
        "version": data_version,
        "last_updated": cached_data["last_updated"]
    }


# ===== EXCEL FILE OPEN/CLOSE =====
class ExcelFileRequest(BaseModel):
    file_path: str


@app.post("/api/open-excel")
async def open_excel(request: ExcelFileRequest):
    """Open an Excel file with the system default application."""
    abs_path = os.path.abspath(request.file_path)
    allowed_paths = [os.path.abspath(fp) for fp in FILE_PATHS]

    if abs_path not in allowed_paths:
        raise HTTPException(status_code=403, detail="File not in allowed paths")

    if not os.path.exists(abs_path):
        raise HTTPException(status_code=404, detail="File not found")

    # If already tracked as open, return current state
    if abs_path in excel_processes:
        proc = excel_processes[abs_path]
        if proc.poll() is None:
            return {"status": "already_open", "file_path": abs_path}
        else:
            # Process ended, clean up
            del excel_processes[abs_path]

    try:
        if sys.platform == 'win32':
            proc = subprocess.Popen(['cmd', '/c', 'start', '', abs_path])
        elif sys.platform == 'darwin':
            proc = subprocess.Popen(['open', abs_path])
        else:
            proc = subprocess.Popen(['xdg-open', abs_path])

        excel_processes[abs_path] = proc
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Opened Excel file: {os.path.basename(abs_path)}")
        return {"status": "opened", "file_path": abs_path}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to open file: {str(e)}")


@app.post("/api/close-excel")
async def close_excel(request: ExcelFileRequest):
    """Close a previously opened Excel file."""
    abs_path = os.path.abspath(request.file_path)
    proc = excel_processes.get(abs_path)

    if not proc:
        return {"status": "not_tracked", "message": "No tracked process for this file"}

    try:
        if proc.poll() is not None:
            # Process already ended
            del excel_processes[abs_path]
            return {"status": "already_closed"}

        if sys.platform == 'win32':
            # On Windows, kill the process tree
            subprocess.run(
                ['taskkill', '/F', '/T', '/PID', str(proc.pid)],
                capture_output=True, timeout=5
            )
        else:
            proc.terminate()
            try:
                proc.wait(timeout=5)
            except subprocess.TimeoutExpired:
                proc.kill()

        del excel_processes[abs_path]
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Closed Excel file: {os.path.basename(abs_path)}")
        return {"status": "closed"}

    except Exception as e:
        excel_processes.pop(abs_path, None)
        return {"status": "closed", "message": f"Process cleanup attempted: {str(e)}"}


@app.get("/api/excel-status")
async def excel_status():
    """Return the current open/close status of all tracked Excel files."""
    status = {}
    to_remove = []
    for path, proc in excel_processes.items():
        if proc.poll() is None:
            status[path] = "open"
        else:
            to_remove.append(path)
    for path in to_remove:
        del excel_processes[path]
    return {"processes": status}


# ===== TASK UPDATE MODEL =====
class TaskUpdate(BaseModel):
    file_path: str
    sheet_name: str
    row_index: int
    task_name: str  # For validation that row hasn't shifted
    updates: Dict[str, Any]  # {column_name: new_value}
    new_columns: Optional[Dict[str, Any]] = None  # New columns to create


@app.post("/api/save-task")
async def save_task(update: TaskUpdate):
    """Save task changes back to the Excel file."""
    global write_in_progress

    try:
        # Validate file exists and is in allowed paths
        abs_path = os.path.abspath(update.file_path)
        allowed_paths = [os.path.abspath(fp) for fp in FILE_PATHS]

        if abs_path not in allowed_paths:
            raise HTTPException(status_code=403, detail="File not in allowed paths")

        if not os.path.exists(abs_path):
            raise HTTPException(status_code=404, detail="File not found")

        # Set write lock to prevent watchdog reload
        with write_lock:
            write_in_progress = True

        try:
            # Read entire Excel file (all sheets) using safe read to handle open files
            excel_data = safe_read_excel(abs_path, sheet_name=None, engine='openpyxl')

            if update.sheet_name not in excel_data:
                raise HTTPException(status_code=404, detail="Sheet not found")

            df = excel_data[update.sheet_name]

            # Validate row index
            if update.row_index < 0 or update.row_index >= len(df):
                raise HTTPException(status_code=400, detail="Invalid row index")

            # Validate task name hasn't shifted
            task_name_col = df.columns[0]
            current_task_name = str(df.iloc[update.row_index][task_name_col])
            if current_task_name != update.task_name:
                raise HTTPException(
                    status_code=409,
                    detail=f"Row position changed. Expected '{update.task_name}' but found '{current_task_name}'. Please refresh and try again."
                )

            # Apply updates to existing columns
            for col, value in update.updates.items():
                if col in df.columns:
                    # Convert empty string to None for proper Excel handling
                    df.at[update.row_index, col] = value if value != '' else None

            # Handle new columns
            if update.new_columns:
                for col, value in update.new_columns.items():
                    if col not in df.columns:
                        # Add new column with None for all rows
                        df[col] = None
                    # Set value for this row
                    df.at[update.row_index, col] = value if value != '' else None

            # Update the sheet in our data structure
            excel_data[update.sheet_name] = df

            # Read original formatting before overwriting (using shared access to handle open files)
            original_col_widths = {}
            original_col_formats = {}  # Store number formats per column
            original_tab_colors = {}  # Store sheet tab colors
            original_book_views = None  # Store workbook view settings (includes tab bar ratio)
            try:
                file_bytes = read_file_with_shared_access(abs_path)
                temp_wb = load_workbook(io.BytesIO(file_bytes))

                # Save workbook view settings (includes tab bar width/ratio)
                if temp_wb.views:
                    original_book_views = temp_wb.views

                for sheet_name in temp_wb.sheetnames:
                    original_col_widths[sheet_name] = {}
                    original_col_formats[sheet_name] = {}
                    ws = temp_wb[sheet_name]

                    # Get sheet tab color
                    if ws.sheet_properties.tabColor:
                        original_tab_colors[sheet_name] = ws.sheet_properties.tabColor

                    # Get column widths
                    for col_letter, dim in ws.column_dimensions.items():
                        if dim.width:
                            original_col_widths[sheet_name][col_letter] = dim.width
                    # Get number formats from the first data row (row 2, after header)
                    if ws.max_row >= 2:
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=2, column=col_idx)
                            if cell.number_format and cell.number_format != 'General':
                                from openpyxl.utils import get_column_letter
                                col_letter = get_column_letter(col_idx)
                                original_col_formats[sheet_name][col_letter] = cell.number_format
                temp_wb.close()
            except Exception:
                pass  # If we can't read formatting, continue without it

            # Write back to Excel file
            try:
                with pd.ExcelWriter(abs_path, engine='openpyxl', mode='w') as writer:
                    for sname, sheet_df in excel_data.items():
                        sheet_df.to_excel(writer, sheet_name=sname, index=False)

                    # Restore workbook view settings (tab bar width/ratio)
                    if original_book_views:
                        writer.book.views = original_book_views

                    # Restore original column widths, formats, and tab colors
                    for sname in writer.sheets:
                        ws = writer.sheets[sname]

                        # Restore sheet tab color
                        if sname in original_tab_colors:
                            ws.sheet_properties.tabColor = original_tab_colors[sname]

                        # Restore column widths
                        if sname in original_col_widths:
                            for col_letter, width in original_col_widths[sname].items():
                                ws.column_dimensions[col_letter].width = width
                        # Restore number formats for all data rows
                        if sname in original_col_formats:
                            for col_letter, num_format in original_col_formats[sname].items():
                                from openpyxl.utils import column_index_from_string
                                col_idx = column_index_from_string(col_letter)
                                # Apply format to all data rows (skip header row 1)
                                for row_idx in range(2, ws.max_row + 1):
                                    ws.cell(row=row_idx, column=col_idx).number_format = num_format
            except PermissionError:
                raise HTTPException(
                    status_code=423,
                    detail="Cannot save: The Excel file is open in another program. Please close Excel and try again."
                )

            print(f"[{datetime.now().strftime('%H:%M:%S')}] Saved changes to {os.path.basename(abs_path)}, sheet '{update.sheet_name}', row {update.row_index}")

            # Small delay to let file system settle
            time.sleep(0.5)

        finally:
            # Release write lock
            with write_lock:
                write_in_progress = False

        # Trigger manual reload after write
        reload_data()

        return {"status": "success", "message": "Task updated successfully"}

    except HTTPException:
        raise
    except Exception as e:
        with write_lock:
            write_in_progress = False
        raise HTTPException(status_code=500, detail=f"Error saving task: {str(e)}")


@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request):
    # Use cached data instead of loading fresh
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "all_sheets_data": cached_data["all_sheets_data"],
            "sheet_names": cached_data["sheet_names"],
            "data_version": data_version
        }
    )


@app.on_event("startup")
async def startup_event():
    """Load initial data and start file watcher on app startup."""
    # Load initial data
    reload_data()
    # Start file watcher in background thread
    threading.Thread(target=start_file_watcher, daemon=True).start()


@app.on_event("shutdown")
async def shutdown_event():
    """Stop file watcher on shutdown."""
    try:
        observer.stop()
        observer.join()
    except Exception:
        pass


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8889, reload=False)
