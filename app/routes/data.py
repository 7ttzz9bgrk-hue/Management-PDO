import io
import logging
import os
import tempfile
import time
from fastapi import APIRouter, HTTPException
from openpyxl import load_workbook
import pandas as pd

from app.config import FILE_PATHS
from app.models import TaskUpdate
from app.services.excel_io import read_file_with_shared_access, safe_read_excel
from app.services.data_loader import reload_data
from app.services.path_guard import is_allowed_path, normalize_path
import app.state as state

router = APIRouter()
logger = logging.getLogger(__name__)


@router.get("/data")
async def get_data():
    """Fetch the latest cached data."""
    return {
        "all_sheets_data": state.cached_data["all_sheets_data"],
        "sheet_names": state.cached_data["sheet_names"],
        "version": state.data_version,
        "last_updated": state.cached_data["last_updated"],
    }


@router.post("/save-task")
async def save_task(update: TaskUpdate):
    """Save task changes back to the Excel file."""
    try:
        abs_path = normalize_path(update.file_path)

        if update.new_columns is None:
            update.new_columns = {}

        if not update.updates and not update.new_columns:
            raise HTTPException(status_code=400, detail="No changes provided")

        if not is_allowed_path(abs_path, FILE_PATHS):
            raise HTTPException(status_code=403, detail="File not in allowed paths")

        _, ext = os.path.splitext(abs_path)
        if ext.lower() not in {".xlsx", ".xlsm", ".xls"}:
            raise HTTPException(status_code=400, detail="Only Excel files are supported")

        if not os.path.isfile(abs_path):
            raise HTTPException(status_code=404, detail="File not found")

        with state.write_lock:
            state.write_in_progress = True

        try:
            excel_data = safe_read_excel(abs_path, sheet_name=None, engine="openpyxl")

            if update.sheet_name not in excel_data:
                raise HTTPException(status_code=404, detail="Sheet not found")

            df = excel_data[update.sheet_name]

            if update.row_index < 0 or update.row_index >= len(df):
                raise HTTPException(status_code=400, detail="Invalid row index")

            task_name_col = df.columns[0]
            current_task_name = str(df.iloc[update.row_index][task_name_col])
            if current_task_name != update.task_name:
                raise HTTPException(
                    status_code=409,
                    detail=f"Row position changed. Expected '{update.task_name}' but found '{current_task_name}'. Please refresh and try again.",
                )

            invalid_update_columns = [
                str(col) for col in update.updates
                if str(col).strip() == ""
            ]
            invalid_new_columns = [
                str(col) for col in update.new_columns
                if str(col).strip() == ""
            ]
            if invalid_update_columns or invalid_new_columns:
                raise HTTPException(
                    status_code=400,
                    detail="Column names cannot be blank",
                )

            unknown_columns = [col for col in update.updates if col not in df.columns]
            if unknown_columns:
                raise HTTPException(
                    status_code=400,
                    detail=f"Unknown column(s): {', '.join(unknown_columns)}",
                )

            overlapping_columns = set(update.updates).intersection(update.new_columns)
            if overlapping_columns:
                overlap_list = ", ".join(sorted(overlapping_columns))
                raise HTTPException(
                    status_code=400,
                    detail=f"Columns cannot appear in both updates and new_columns: {overlap_list}",
                )

            for col, value in update.updates.items():
                df.at[update.row_index, col] = value if value != "" else None

            if update.new_columns:
                for col, value in update.new_columns.items():
                    if col not in df.columns:
                        df[col] = None
                    df.at[update.row_index, col] = value if value != "" else None

            excel_data[update.sheet_name] = df

            original_col_widths, original_col_formats, original_tab_colors, original_book_views = _read_formatting(abs_path)

            try:
                _write_excel(abs_path, excel_data, original_col_widths, original_col_formats, original_tab_colors, original_book_views)
            except (PermissionError, OSError) as err:
                if isinstance(err, PermissionError) or getattr(err, "errno", None) == 13:
                    raise HTTPException(
                        status_code=423,
                        detail="Cannot save: The Excel file is open in another program. Please close Excel and try again.",
                    )
                raise

            logger.info(
                "Saved changes to %s, sheet '%s', row %d",
                os.path.basename(abs_path),
                update.sheet_name,
                update.row_index,
            )
            time.sleep(0.5)

        finally:
            with state.write_lock:
                state.write_in_progress = False

        reload_data()
        return {"status": "success", "message": "Task updated successfully"}

    except HTTPException:
        raise
    except Exception as e:
        with state.write_lock:
            state.write_in_progress = False
        raise HTTPException(status_code=500, detail=f"Error saving task: {str(e)}")


def _read_formatting(abs_path):
    """Read original Excel formatting before overwriting."""
    original_col_widths = {}
    original_col_formats = {}
    original_tab_colors = {}
    original_book_views = None

    try:
        file_bytes = read_file_with_shared_access(abs_path)
        temp_wb = load_workbook(io.BytesIO(file_bytes))

        if temp_wb.views:
            original_book_views = temp_wb.views

        for sheet_name in temp_wb.sheetnames:
            original_col_widths[sheet_name] = {}
            original_col_formats[sheet_name] = {}
            ws = temp_wb[sheet_name]

            if ws.sheet_properties.tabColor:
                original_tab_colors[sheet_name] = ws.sheet_properties.tabColor

            for col_letter, dim in ws.column_dimensions.items():
                if dim.width:
                    original_col_widths[sheet_name][col_letter] = dim.width

            if ws.max_row >= 2:
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=2, column=col_idx)
                    if cell.number_format and cell.number_format != "General":
                        from openpyxl.utils import get_column_letter
                        col_letter = get_column_letter(col_idx)
                        original_col_formats[sheet_name][col_letter] = cell.number_format

        temp_wb.close()
    except Exception as exc:
        logger.warning("Could not preserve original workbook formatting: %s", exc)

    return original_col_widths, original_col_formats, original_tab_colors, original_book_views


def _write_excel(abs_path, excel_data, col_widths, col_formats, tab_colors, book_views):
    """Write Excel data back to file atomically, restoring formatting.

    Writes to a temp file in the same directory first, then replaces the
    original via rename. This prevents data loss if the write fails midway.
    """
    dir_name = os.path.dirname(abs_path)
    tmp_fd, tmp_path = tempfile.mkstemp(dir=dir_name, suffix=".tmp.xlsx")
    try:
        os.close(tmp_fd)
        with pd.ExcelWriter(tmp_path, engine="openpyxl", mode="w") as writer:
            for sname, sheet_df in excel_data.items():
                sheet_df.to_excel(writer, sheet_name=sname, index=False)

            if book_views:
                writer.book.views = book_views

            for sname in writer.sheets:
                ws = writer.sheets[sname]

                if sname in tab_colors:
                    ws.sheet_properties.tabColor = tab_colors[sname]

                if sname in col_widths:
                    for col_letter, width in col_widths[sname].items():
                        ws.column_dimensions[col_letter].width = width

                if sname in col_formats:
                    for col_letter, num_format in col_formats[sname].items():
                        from openpyxl.utils import column_index_from_string
                        col_idx = column_index_from_string(col_letter)
                        for row_idx in range(2, ws.max_row + 1):
                            ws.cell(row=row_idx, column=col_idx).number_format = num_format

        os.replace(tmp_path, abs_path)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise
